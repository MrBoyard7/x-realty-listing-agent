from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from realty_agent.models import EXCEL_COLUMNS, ExtractedListing, ListingStatus
from realty_agent.storage.excel_writer import (
    ensure_workbook,
    insert_error_row,
    insert_row_at_top,
    read_active_rows,
    set_row_status,
)

PHX = ZoneInfo("America/Phoenix")


def make_listing(post_id="1", price=300000) -> ExtractedListing:
    return ExtractedListing(
        post_id=post_id,
        post_date_time=datetime(2026, 6, 1, 9, 0, tzinfo=PHX),
        post_url=f"https://x.com/x/status/{post_id}",
        notes="test note",
        address="123 Test St",
        city="Phoenix",
        state="AZ",
        beds=3,
        baths=2,
        square_feet=1400,
        price=price,
    )


def test_ensure_workbook_creates_header_row(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    ensure_workbook(wb_path)
    wb = load_workbook(wb_path)
    ws = wb["Listings"]
    header = [c.value for c in ws[1]]
    assert header == EXCEL_COLUMNS


def test_insert_row_at_top_keeps_descending_order(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    ensure_workbook(wb_path)

    insert_row_at_top(wb_path, make_listing(post_id="1"))
    insert_row_at_top(wb_path, make_listing(post_id="2"))

    wb = load_workbook(wb_path)
    ws = wb["Listings"]
    # Most recently inserted listing should now be in row 2
    assert ws.cell(row=2, column=EXCEL_COLUMNS.index("X Post ID") + 1).value == "2"
    assert ws.cell(row=3, column=EXCEL_COLUMNS.index("X Post ID") + 1).value == "1"


def test_new_rows_default_to_status_new(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    ensure_workbook(wb_path)
    insert_row_at_top(wb_path, make_listing())
    active = read_active_rows(wb_path)
    assert len(active) == 1
    assert active[0]["Status"] == ListingStatus.NEW.value


def test_set_row_status_archives_row(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    ensure_workbook(wb_path)
    insert_row_at_top(wb_path, make_listing())
    set_row_status(wb_path, 2, ListingStatus.ARCHIVE)
    active = read_active_rows(wb_path)
    assert active == []


def test_insert_error_row(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    ensure_workbook(wb_path)
    insert_error_row(wb_path, datetime(2026, 6, 1, 9, 0, tzinfo=PHX), "X API timeout")
    wb = load_workbook(wb_path)
    ws = wb["Listings"]
    status_col = EXCEL_COLUMNS.index("Status") + 1
    notes_col = EXCEL_COLUMNS.index("Notes") + 1
    assert ws.cell(row=2, column=status_col).value == ListingStatus.ERROR.value
    assert ws.cell(row=2, column=notes_col).value == "X API timeout"
