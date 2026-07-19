from pathlib import Path

from openpyxl import load_workbook

from realty_agent.models import EXCEL_COLUMNS, ListingStatus
from realty_agent.storage.excel_writer import ensure_workbook
from realty_agent.sync.delta_sync import run_delta_sync
from realty_agent.x_client.mock_client import MockXClient


def test_full_pipeline_with_sample_data(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    state_path = tmp_path / "state.json"
    ensure_workbook(wb_path)

    client = MockXClient.with_sample_data()
    result = run_delta_sync(client, "test_wholesale_deals", wb_path, state_path)

    # 5 sample posts: 1001 (new property), 1002 (new property), 1003
    # (repost of 1001 w/ price change -> new version row inserted, 1001's
    # row archived), 1004 (exact repost of 1002 -> ignored), 1005
    # (unstructured, no address -> new property row with blank fields).
    # Final sheet: 4 rows total, 3 with Status=New (1002, 1003, 1005) and
    # 1 with Status=Archive (1001).
    assert result.posts_seen == 5
    assert result.duplicates_ignored == 1
    assert result.rows_archived == 1
    assert result.rows_inserted == 4

    wb = load_workbook(wb_path)
    ws = wb["Listings"]
    header = [c.value for c in ws[1]]
    assert header == EXCEL_COLUMNS

    statuses = [
        ws.cell(row=r, column=EXCEL_COLUMNS.index("Status") + 1).value
        for r in range(2, ws.max_row + 1)
    ]
    assert statuses.count(ListingStatus.ARCHIVE.value) == 1
    assert statuses.count(ListingStatus.NEW.value) == 3


def test_delta_sync_only_processes_new_posts_on_second_run(tmp_path: Path):
    wb_path = tmp_path / "listings.xlsx"
    state_path = tmp_path / "state.json"
    ensure_workbook(wb_path)

    client = MockXClient.with_sample_data()
    first = run_delta_sync(client, "test_wholesale_deals", wb_path, state_path)
    assert first.posts_seen == 5

    second = run_delta_sync(client, "test_wholesale_deals", wb_path, state_path)
    assert second.posts_seen == 0
