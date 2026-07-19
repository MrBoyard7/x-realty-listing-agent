"""Read/write helpers for the listings workbook.

Operates on a local ``.xlsx`` file. In production that file is the one
synced to/from OneDrive by :mod:`realty_agent.storage.onedrive_client`
(download -> mutate locally with these helpers -> upload), which keeps
the row-insertion logic simple, fast, and independent of the Graph API.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from realty_agent.models import EXCEL_COLUMNS, ExtractedListing, ListingStatus


def ensure_workbook(path: Path) -> None:
    """Create the workbook with a header row if it does not exist yet."""
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    ws.append(EXCEL_COLUMNS)
    wb.save(path)


def _row_to_dict(ws: Worksheet, row_index: int) -> dict:
    return {
        col: ws.cell(row=row_index, column=idx + 1).value for idx, col in enumerate(EXCEL_COLUMNS)
    }


def read_active_rows(path: Path) -> List[dict]:
    """Return only rows whose Status is New/Like/blank, as plain dicts.

    Archived rows are skipped entirely so they are never re-sent to the
    AI model or re-evaluated for matching, per the cost-optimization
    requirement.
    """
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb["Listings"]
    rows = []
    for row_index in range(2, ws.max_row + 1):
        row = _row_to_dict(ws, row_index)
        if row.get("Status") in (ListingStatus.NEW.value, ListingStatus.LIKE.value, None, ""):
            row["_row_index"] = row_index
            rows.append(row)
    return rows


def _listing_to_row_values(listing: ExtractedListing, status: ListingStatus) -> list:
    local_dt = listing.post_date_time
    return [
        local_dt.strftime("%Y-%m-%d %H:%M:%S"),  # SQL Server-compatible datetime
        local_dt.strftime("%m/%d/%Y"),
        local_dt.strftime("%H:%M:%S"),
        listing.post_id,
        listing.post_url,
        status.value,
        listing.address,
        listing.city,
        listing.state,
        listing.zip_code,
        listing.county,
        listing.parcel_number,
        listing.beds,
        listing.baths,
        listing.square_feet,
        listing.price,
        listing.arv,
        listing.age_of_roof,
        listing.age_of_ac,
        listing.google_docs_url,
        listing.zillow_url,
        listing.redfin_url,
        listing.county_assessor_url,
        listing.notes,
    ]


def insert_row_at_top(
    path: Path,
    listing: ExtractedListing,
    status: ListingStatus = ListingStatus.NEW,
) -> None:
    """Insert ``listing`` at row 2, pushing existing data rows down.

    Keeps the sheet in descending order by post date/time as required.
    """
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb["Listings"]
    ws.insert_rows(2)
    values = _listing_to_row_values(listing, status)
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    wb.save(path)


def set_row_status(path: Path, row_index: int, status: ListingStatus) -> None:
    """Update the Status cell for an existing row (e.g. archive it)."""
    wb = load_workbook(path)
    ws = wb["Listings"]
    status_col = EXCEL_COLUMNS.index("Status") + 1
    ws.cell(row=row_index, column=status_col, value=status.value)
    wb.save(path)


def insert_error_row(path: Path, error_dt: datetime, message: str) -> None:
    """Insert an error row at row 2 per the spec's error-handling rules."""
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb["Listings"]
    ws.insert_rows(2)
    values: List[Optional[str]] = [None] * len(EXCEL_COLUMNS)
    col = {name: i for i, name in enumerate(EXCEL_COLUMNS)}
    values[col["Post Date/Time"]] = error_dt.strftime("%Y-%m-%d %H:%M:%S")
    values[col["Post Date"]] = error_dt.strftime("%m/%d/%Y")
    values[col["Post Time"]] = error_dt.strftime("%H:%M:%S")
    values[col["Status"]] = ListingStatus.ERROR.value
    values[col["Notes"]] = message
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    wb.save(path)
